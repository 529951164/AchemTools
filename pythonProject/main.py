
import openpyxl
import csv
import os
import json
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)

def readCSV():
    with open('./result_K1-3_K2-1.csv', 'r', encoding='Shift-JIS') as f:
        read = csv.reader(f)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "test"
        for line in read:
            print(line)
            ws.append(line)

    wb.save('example.xlsx')
    print('done')

def readExcel():
    from openpyxl import Workbook, load_workbook
    wb = load_workbook('./result_K1-3_K2-1.csv')
    ws = wb.active
    for row in ws:
        for cell in row:
            ws.append(cell)
            print(cell)

    wb.save('example.xlsx')
    print('done')

def add_chart(path, file_name):
    wb = Workbook()
    ws = wb.active
    if not os.path.exists('excel'):
        os.makedirs('excel')

    with open(path, 'r', encoding='Shift-JIS') as f:
        read = csv.reader(f)
        for line in read:
            ws.append(line)

        for row in ws:
            for cell in row:
                # cell.number_format = '0.00'
                if cell.value != None:
                    try:
                        print(float(cell.value))
                        cell.value = float(cell.value)
                    except Exception as e:
                        print('str:' + cell.value)


    c1 = LineChart()
    # c1.title = "k1=3,k2=1"
    c1.style = 13
    c1.y_axis.title = '生成物浓度'
    c1.x_axis.title = '时间'
    data = Reference(ws, min_col=5, min_row=3, max_col=6, max_row=41)
    cats = Reference(ws, min_col=2, min_row=4, max_row=41)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    # Style the lines
    s1 = c1.series[0]
    s1.marker.symbol = "triangle"
    s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
    s1.graphicalProperties.line.noFill = True
    s2 = c1.series[1]
    s2.graphicalProperties.line.solidFill = "00AAAA"
    s2.graphicalProperties.line.dashStyle = "sysDot"
    s2.graphicalProperties.line.width = 100050  # width in EMUs
    ws.add_chart(c1, "i10")
    wb.save("excel/" + file_name + '.xlsx')


def file_name(file_dir):
    allJson = '[';
    for root, dirs, files in os.walk(file_dir):
        print('files:', files)  # 当前路径下所有非目录子文件
        for file in files:
            name_str = file.replace('.csv','')
            add_chart('./csv/' + file, name_str)
            allJson += export_json(name_str) + ","

    allJson += ']'
    allJson = allJson.replace(',]', ']')
    allJson = allJson.replace('    ', '')
    json_file = open('./database/all.json', 'w+', encoding='utf-8')
    json_file.write(allJson)
    json_file.close()


def export_json(file_name):
    if not os.path.exists('database'):
        os.makedirs('database')
    result_str = '{'
    # json_file = open('./database/' + file_name + '.json', 'w+', encoding='utf-8')
    csv_file = open('./csv/' + file_name + '.csv', 'r', encoding='utf-8')
    result_str += f'"{file_name}":'
    # 读取文件第3行不读取换行符作为json文件里每个数据的键值
    csv_file.readline()
    csv_file.readline()
    keys = csv_file.readline().strip('\n').split(',')
    # json_file.write('[\n')
    result_str += '['
    flag = 0
    line = csv_file.readline()
    while line:

        # 字符串列表转化为数字列表
        values = line.strip('\n').split(',')
        values_temp = map(eval, values[0:6])
        values_else = list(values_temp)
        # values_else.append(values[2])
        # 用zip()函数将两个列表形成映射关系，创建字典
        dic_temp = dict(zip(keys, values_else))
        # 将字典转化为字符串且带有缩进
        # flag用于判断json文件中 "," 和换行的添加位置
        json_str = json.dumps(dic_temp, indent=4)
        json_str = json_str.replace("\n", "", 100)
        json_str += ','
        # json_file.write(json_str)
        result_str += json_str

        line = csv_file.readline()

    # json_file.write(']')
    result_str += ']}'
    csv_file.close()
    # json_file.close()
    result_str = result_str.replace(',]', ']')
    return result_str


if __name__ == '__main__':
    file_name('csv')
    # export_json()
